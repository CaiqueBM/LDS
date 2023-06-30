from flask import (
    Flask,
    session,
    render_template,
    make_response,
    request,
    redirect,
    url_for,
    flash,
    send_file,
)
from numpy._typing import _80Bit
import pandas as pd
import datetime
import sqlite3
import csv
from openpyxl import Workbook, load_workbook


# from babkuppy import para_avaliacao
from cria_tabelas import gerar
from flask_bootstrap import Bootstrap
import os
import re
import shutil
import difflib
import xlwings as xw
import ast

app = Flask(__name__)
app.static_folder = "static"
app.secret_key = "2@2"
bootstrap = Bootstrap(app)

gerar()
df_tabela = pd.DataFrame
df_arquivo = pd.DataFrame(columns=["nome", "projeto", "titulo"])
atualizar = "intermedio"
pasta_destino = ""
linha_selecionada = ""
df_projeto = pd.DataFrame(columns=["projeto", "abreviacao", "descricao"])
novo_caminho = ""
control = ""
responsavel_status = ""
df_selecionado = pd.DataFrame
aprovado_exibido = False
mudar_status = ""
abreviacao = ""
data_atualizada = ""
diretorio_raiz = os.path.join("/", "media", "hdfs", "Engenharia", "Projetos")
diretorio_default = os.path.join(
    "/", "media", "hdfs", "Engenharia", "Projetos", "Sistema LDS", "Modelos de Arquivos"
)
caminho_padrao = os.path.join(
    "/", "media", "hdfs", "Engenharia", "Projetos", "Sistema LDS", "GRD E LD padrao"
)
pasta_padrao_projeto = os.path.join(
    "/", "media", "hdfs", "Engenharia", "Projetos", "0000 - Novo Projeto"
)

#diretorio_raiz = r"""C:\Users\lanch\Desktop\Projeto"""
#diretorio_default = r"""C:\Users\lanch\Desktop\Default"""
#caminho_padrao = r"""C:\Users\lanch\Desktop\modeloGRD"""
#pasta_padrao_projeto = r"""C:\Users\lanch\Desktop\1 - Padrao"""



@app.route("/")
def index():
    return render_template("login.html")


# Rota login de usuario


@app.route("/login", methods=["POST"])
def login():
    global abreviacao
    username = request.form.get("username")
    password = request.form.get("password")

    with open("users.csv", "r", encoding="utf-8") as file:
        for line in file:
            fields = line.strip().split(";")
            if fields[0] == username and fields[1] == password:
                now = datetime.datetime.now()
                login_time = now.strftime("%d/%m/%Y %H:%M:%S")
                conn = sqlite3.connect("database.db")
                c = conn.cursor()
                c.execute(
                    "INSERT INTO log (username, login_time) VALUES (?, ?)",
                    (username, login_time),
                )
                conn.commit()
                conn.close()
                session["username"] = username
                session["login_time"] = login_time
                session["abreviacao"] = fields[2]
                abreviacao = session["abreviacao"]
                return redirect("/user")

    flash("Nome de usuário ou senha incorretos")
    return render_template("login.html")


# Rota para sair da sessão


@app.route("/logout")
def logout():
    session.pop("username", None)  # Remover nome do usuário da sessão
    return redirect(url_for("index"))


# ----------------- Rota da pagina usuario, todas acoes serao dadas aqui---------------


@app.route("/user")
def user():
    global df_tabela

    if "username" in session:
        # recuperando o nome de usuário da sessão
        username = session["username"]
        login_time = session["login_time"]
        conn = sqlite3.connect("database.db")
        df_tabela = pd.read_sql_query(
            f"SELECT DISTINCT projeto FROM arquivos WHERE (responsavel = '{username}' OR aprovador = '{username}')",
            conn,
        )

        df_tabela["projeto"] = df_tabela["projeto"].apply(
            lambda x: f"<a href='/user/{x}'>{x}</a>"
        )

        # Feche a conexão com o banco de dados
        conn.close()
        tabela_projetos = df_tabela.to_html(
            classes="table table-striped table-user", escape=False, index=False
        )
        return render_template(
            "user.html",
            username=username,
            login_time=login_time,
            tabela_projetos=tabela_projetos,
        )
    else:
        # redirecionando para a página de login se o nome de usuário não estiver na sessão
        return redirect("/")


# Rota para abrir os documentos de um determinado projeto do usuario
@app.route("/user/<projeto>", methods=["GET", "POST"])
def user_projetos(projeto):
    if "username" in session:
        global df_tabela
        global pasta_destino
        global atualizar
        global control
        global responsavel_status
        global df_selecionado
        global aprovado_exibido
        global mudar_status
        global abreviacao
        global data_atualizada

        username = session["username"]
        login_time = session["login_time"]
        abreviacao = session["abreviacao"]

        conn = sqlite3.connect("database.db")

        df_tabela = pd.read_sql_query(
            f"SELECT * FROM arquivos WHERE projeto=? AND (responsavel=? OR aprovador=?)",
            conn,
            params=(projeto, username, username),
        )

        conn.close()

        if control == "atualizar":
            atualizar = "atualizar"
        elif control == "renomear":
            atualizar = "renomear"
        elif control == "gerar_grd":
            atualizar = "gerar_grd"
        else:
            atualizar = "intermedio"

        nome_pasta = False
        doc = [tuple(row) for row in df_tabela.values]
        # obtém a URL da página anterior
        status_list = df_tabela["status"].unique()

        if not df_selecionado.empty:
            doc_selecionado = [tuple(row) for row in df_selecionado.values]
            for item in doc_selecionado:
                if item[9] is not None and username in item[9]:
                    aprovado_exibido = False
                    break
                else:
                    aprovado_exibido = True

            return render_template(
                "user_projetos.html",
                username=username,
                login_time=login_time,
                projeto=projeto,
                status_list=status_list,
                doc=doc,
                nome_pasta=nome_pasta,
                pasta_destino=pasta_destino,
                atualizar=atualizar,
                responsavel_status=responsavel_status,
                doc_selecionado=doc_selecionado,
                aprovado_exibido=aprovado_exibido,
                abreviacao=abreviacao,
                data_atualizada=data_atualizada,
            )

        return render_template(
            "user_projetos.html",
            username=username,
            login_time=login_time,
            projeto=projeto,
            status_list=status_list,
            doc=doc,
            nome_pasta=nome_pasta,
            pasta_destino=pasta_destino,
            atualizar=atualizar,
            responsavel_status=responsavel_status,
            abreviacao=abreviacao,
            data_atualizada=data_atualizada,
        )


@app.route("/projetos", methods=["GET"])
def projetos():
    if "username" in session:
        username = session["username"]
        login_time = session["login_time"]
        global df_tabela
        global diretorio_raiz
        df_link = pd.DataFrame(columns=["projeto", "numero"])

        for pasta in os.listdir(diretorio_raiz):
            if os.path.isdir(os.path.join(diretorio_raiz, pasta)):
                if re.search(r"^(00+|[^0-9].*|[^0-9].*-)", pasta):
                    continue
                caminho_projeto = os.path.join(diretorio_raiz, pasta)


                numero_personalizado = re.search(r"Projetos/(\d+)\s*-", caminho_projeto)
                if numero_personalizado:
                    numero_projeto = numero_personalizado.group(1)

                #Buscar nome do projeto
                projeto_arquivo = re.search(
                    r"\d...([A-Za-z\s]+[\w-]+)", caminho_projeto
                )
                if projeto_arquivo:
                    projeto = projeto_arquivo.group(1)
                else:
                    projeto = ""
                    numero_projeto = ""
                df_link.loc[len(df_link)] = [projeto, numero_projeto]
        df_link = df_link.sort_values(by='numero', ascending=False)
        df_link = df_link.drop('numero', axis=1)

        # Adiciona um link no nome dos projetos na coluna "projeto" da tabela
        df_link["projeto"] = (
            "<a href='/projetos/"
            + df_link["projeto"]
            + "'>"
            + df_link["projeto"]
            + "</a>"
        )

        # Renderiza a página "projetos.html" e passa os dados da tabela "arquivos" para a variável "tabela"
        tabela = df_link.to_html(
            classes="table table-striped table-user",
            escape=False,
            index=False,
            table_id="myTable",
        )

        return render_template(
            "projetos.html",
            username=username,
            login_time=login_time,
            tabela=tabela,
        )
    else:
        # redirecionando para a página de login se o nome de usuário não estiver na sessão
        return redirect("/")


# Rota para abrir os documentos de um determinado projeto
@app.route("/projetos/<projeto>")
def documentos(projeto):
    if "username" in session:
        global df_tabela
        global diretorio_default
        global df_arquivos

        username = session["username"]
        login_time = session["login_time"]

        # Diretorio dos arquivos Default para criar para começar um projeto
        arquivos_na_pasta = [arquivo for arquivo in os.listdir(diretorio_default)]

        conn = sqlite3.connect("database.db")

        df_tabela = pd.read_sql_query(
            f'''SELECT * FROM arquivos WHERE projeto="{(projeto)}"''',
            conn,
        )
        documentos = [tuple(row) for row in df_tabela.values]

        status_list = df_tabela["status"].unique()

        disc = []
        sub = []

        with open("disc.csv", "r", encoding="utf-8") as arquivo:
            for line in arquivo:
                fields = line.strip().split(";")
                disc.append(fields[1])

        with open("sub.csv", "r", encoding="utf-8") as arquivo:
            for line in arquivo:
                fields = line.strip().split(";")
                sub.append(fields[1])

        conn.close()
        # df_tabela.drop(df_tabela.index, inplace=True)

        response = make_response(
            render_template(
                "documentos.html",
                username=username,
                login_time=login_time,
                projeto=projeto,
                documentos=documentos,
                status_list=status_list,
                arquivos=arquivos_na_pasta,
                disc=disc,
                sub=sub,
            )
        )
        response.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
        response.headers["Pragma"] = "no-cache"
        response.headers["Expires"] = "0"
        return response


# -------------- Atualizar o status do responsavel, gerar GRD ----------------
@app.route("/atualizar_responsavel", methods=["POST"])
def atualizar_responsavel():
    if "username" in session:
        username = session["username"]
        projeto = request.form["projeto"]
        atualizar_responsavel = request.form.get("mudar_responsavel", None)
        linha_selecionada = request.form.getlist("selecionados")
        linha_selecionada = list(map(int, linha_selecionada))
        tamanho_lista = len(linha_selecionada)

        now = datetime.datetime.now()
        data_hora = now.strftime("%d/%m/%Y %H:%M:%S")

        conn = sqlite3.connect("database.db")
        query = "SELECT * FROM arquivos"
        df_tabela = pd.read_sql_query(query, conn)
        df_selecionado = df_tabela.loc[df_tabela["id"].isin(linha_selecionada)]
        status_atual = df_selecionado["status"].values.tolist()
        caminho_projeto = df_selecionado["caminho"].values.tolist()
        id_documentos = df_selecionado["id"].values.tolist()
        conn.close()

        if tamanho_lista <= 0:  # Retirar erro do botao
            return redirect(
                url_for(
                    "documentos",
                    projeto=projeto,
                    atualizar=atualizar,
                    df_selecionado=df_selecionado,
                )
            )
        else:
            if status_atual[0] == "Criado" or status_atual[0] == "Em Desenvolvimento":
                conn = sqlite3.connect("database.db")
                c = conn.cursor()
                for i in range(tamanho_lista):
                    nome_do_arquivo = os.path.basename(caminho_projeto[i])

                    partes_pasta = caminho_projeto[i].split(
                        "\\"
                    )  # Dividindo o caminho do projeto usando as barras duplas como separadores
                    indice_area_trabalho = partes_pasta.index(
                        "Area de Trabalho"
                    )  # Obtendo o índice da parte "Area de Trabalho" na lista
                    caminho_sem_responsavel = "\\".join(
                        partes_pasta[: indice_area_trabalho + 1]
                    )  # Unindo as partes do caminho até a área de trabalho novamente usando as barras duplas
                    caminho_final = (
                        caminho_sem_responsavel
                        + "\\"
                        + username
                        + "\\"
                        + nome_do_arquivo
                    )  # Adicionando o responsável ao final do caminho

                    shutil.move(caminho_projeto[i], caminho_final)

                    # Atualiza o status do arquivo na base de dados

                    query = f"""UPDATE arquivos SET responsavel = '{username}', caminho = '{caminho_final}', status = 'Em Desenvolvimento' WHERE id = '{id_documentos[i]}' """
                    c.execute(query)
                    conn.commit()

                    c.execute(
                        "INSERT INTO log_tarefas (nome, projeto, status, data_status, responsavel) VALUES (?, ?, ?, ?, ?)",
                        (
                            nome_do_arquivo,
                            projeto,
                            status_atual[0],
                            data_hora,
                            username,
                        ),
                    )
                    conn.commit()

                conn.close()
            else:
                print("0")
            return redirect(
                url_for(
                    "documentos",
                    projeto=projeto,
                    responsavel_status=responsavel_status,
                    atualizar=atualizar,
                    df_selecionado=df_selecionado,
                )
            )


# -------------- Atualizar o status do responsavel, gerar GRD ----------------
@app.route("/intermediador", methods=["POST"])
def intermediador():
    if "username" in session:
        global linha_selecionada
        global atualizar
        global control
        global responsavel_status
        global df_tabela
        global df_selecionado
        global mudar_status

        username = session["username"]
        projeto = request.form["projeto"]
        linha_selecionada = request.form.getlist("selecionados")
        linha_selecionada = list(map(int, linha_selecionada))
        tamanho_lista = len(linha_selecionada)

        if tamanho_lista <= 0:
            return redirect(
                url_for(
                    "user_projetos",
                    projeto=projeto,
                    responsavel_status=responsavel_status,
                    atualizar=atualizar,
                    df_selecionado=df_selecionado,
                )
            )
        else:
            conn = sqlite3.connect("database.db")
            query = "SELECT * FROM arquivos"
            df_tabela = pd.read_sql_query(query, conn)
            df_selecionado = df_tabela.loc[df_tabela["id"].isin(linha_selecionada)]
            status_atual = df_selecionado["status"].values.tolist()
            conn.close()

            responsavel_status = status_atual[0]

            control = "atualizar"

            return redirect(
                url_for(
                    "user_projetos",
                    projeto=projeto,
                    responsavel_status=responsavel_status,
                    atualizar=atualizar,
                    df_selecionado=df_selecionado,
                )
            )


@app.route("/atualizar_status", methods=["GET", "POST"])
def atualizar_status():
    if "username" in session:
        global abreviacao
        global df_tabela
        global diretorio_raiz
        global pasta_destino
        global linha_selecionada
        global atualizar
        global control
        global responsavel_status
        global mudar_status
        global data_atualizada
        global abreviacao

        status = [
            "Criado",
            "Em Desenvolvimento",
            "Para Avaliacao",
            "Para Entrega",
            "Entregue",
        ]

        username = session["username"]
        projeto = request.form["projeto"]
        aprovador = request.form.get("aprovador", None)
        responsavel = request.form.get("nome_responsavel", None)
        status_aprovador = request.form.get("status_aprovador", None)
        status_atualizar = request.form.get("status_atualizar", None)
        now = datetime.datetime.now()
        data_atualizada = now.strftime("%d/%m/%Y")
        data_hora = now.strftime("%d/%m/%Y %H:%M:%S")
        tamanho = len(linha_selecionada)

        conn = sqlite3.connect("database.db")
        query = "SELECT * FROM arquivos"
        df_tabela = pd.read_sql_query(query, conn)
        df_selecionado = df_tabela.loc[df_tabela["id"].isin(linha_selecionada)]
        nome_arq = df_selecionado["nome"].values.tolist()
        caminho_projeto = df_selecionado["caminho"].values.tolist()
        status_atual = df_selecionado["status"].values.tolist()
        id_documentos = df_selecionado["id"].values.tolist()
        conn.close()

        pasta_destino = ""

        if tamanho <= 0:
            control = ""
            linha_selecionada = []
            return redirect(
                url_for(
                    "user_projetos",
                    projeto=projeto,
                    atualizar=atualizar,
                    pasta_destino=pasta_destino,
                )
            )

        if (
            status_atualizar == "atualizar"
            or status_aprovador == "aprovado"
            or status_aprovador == "reprovado"
        ):
            if status_atual[0] == "Criado":
                control = ""
                linha_selecionada = []

                if caminho_projeto is not None:
                    print("-----------------------------------------------")
                    print(caminho_projeto[0])
                    print("-----------------------------------------------")

                    projeto_arquivo = re.search(
                        # r"\d....([A-Za-z\s]+[\w-]+)", caminho_projeto[0]
                        r"(?<=\/Projetos\/).*?(?=\/Arquivos\ do\ Projeto)",
                        caminho_projeto[0],
                    )
                    projeto_arquivo = projeto_arquivo.group(0)

                # Itera sobre os resultados e move cada arquivo para a pasta de destino
                conn = sqlite3.connect("database.db")
                c = conn.cursor()
                for i in range(tamanho):
                    nome_do_arquivo = os.path.basename(caminho_projeto[i])

                    # Atualiza o status do arquivo na base de dados
                    novo_status = status[
                        (status.index(status_atual[i]) + 1) % len(status)
                    ]
                    query = f"""UPDATE arquivos SET status = '{novo_status}', responsavel = '{username}', data_avaliacao = '{data_atualizada}' WHERE id = '{id_documentos[i]}' """
                    c.execute(query)
                    conn.commit()

                    c.execute(
                        "INSERT INTO log_tarefas (nome, projeto, status, data_status, responsavel) VALUES (?, ?, ?, ?, ?)",
                        (
                            nome_do_arquivo,
                            projeto,
                            status_atual[0],
                            data_hora,
                            username,
                        ),
                    )
                    conn.commit()

                conn.close()
                linha_selecionada = []
                df_tabela = df_tabela.dropna()
                df_selecionado = df_selecionado.dropna()
            elif status_atual[0] == "Em Desenvolvimento":
                control = ""
                linha_selecionada = []
                if caminho_projeto is not None:
                    projeto_arquivo = re.search(
                        r"(?<=\/Projetos\/).*?(?=\/Arquivos\ do\ Projeto)",
                        caminho_projeto[0],
                    )
                    projeto_arquivo = projeto_arquivo.group(0)
                    print("-----------------------------------------------")
                    print("projeto_arquivo:", projeto_arquivo)
                    print("-----------------------------------------------")

                pasta_destino = os.path.join(
                    diretorio_raiz,
                    projeto_arquivo,
                    "Arquivos do Projeto",
                    "Para Avaliacao",
                )

                # Itera sobre os resultados e move cada arquivo para a pasta de destino
                conn = sqlite3.connect("database.db")
                c = conn.cursor()
                for i in range(tamanho):
                    nome_do_arquivo = os.path.basename(caminho_projeto[i])
                    novo_caminho = os.path.join(pasta_destino, nome_arq[i])

                    # Atualiza o status do arquivo na base de dados
                    novo_status = status[
                        (status.index(status_atual[i]) + 1) % len(status)
                    ]
                    query = f"""UPDATE arquivos SET caminho='{novo_caminho}', status = '{novo_status}', responsavel = '{username}', data_avaliacao = '{data_atualizada}', aprovador = '{aprovador}' WHERE id = '{id_documentos[i]}' """
                    c.execute(query)

                    c.execute(
                        "INSERT INTO log_tarefas (nome, projeto, status, data_status, responsavel) VALUES (?, ?, ?, ?, ?)",
                        (
                            nome_do_arquivo,
                            projeto,
                            status_atual[0],
                            data_atualizada,
                            username,
                        ),
                    )
                    conn.commit()

                    shutil.move(caminho_projeto[i], pasta_destino)

                conn.close()
                linha_selecionada = []
                df_tabela = df_tabela.dropna()
            elif status_atual[0] == "Para Avaliacao":
                if status_aprovador == "reprovado":
                    control = ""
                    linha_selecionada = []

                    if caminho_projeto is not None:
                        projeto_arquivo = re.search(
                            r"(?<=\/Projetos\/).*?(?=\/Arquivos\ do\ Projeto)",
                            caminho_projeto[0],
                        )
                    projeto_arquivo = projeto_arquivo.group(0)
                    print("-----------------------------------------------")
                    print("projeto_arquivo:", projeto_arquivo)
                    print("-----------------------------------------------")

                    pasta_destino = os.path.join(
                        diretorio_raiz,
                        projeto_arquivo,
                        "Arquivos do Projeto",
                        "Area de Trabalho",
                        responsavel,
                    )

                    # Itera sobre os resultados e move cada arquivo para a pasta de destino
                    conn = sqlite3.connect("database.db")
                    c = conn.cursor()
                    for i in range(tamanho):
                        nome_do_arquivo = os.path.basename(caminho_projeto[i])
                        novo_caminho = os.path.join(pasta_destino, nome_arq[i])

                        # Atualiza o status do arquivo na base de dados
                        novo_status = status[
                            (status.index(status_atual[i]) + 1) % len(status)
                        ]
                        query = f"""UPDATE arquivos SET caminho='{novo_caminho}', status = 'Em Desenvolvimento', responsavel = '{responsavel}', aprovador = '' WHERE id = '{id_documentos[i]}' """
                        c.execute(query)
                        conn.commit()

                        shutil.move(caminho_projeto[i], pasta_destino)

                    linha_selecionada = []
                    df_tabela = df_tabela.dropna()
                elif status_aprovador == "aprovado":
                    control = ""

                    projeto_arquivo = re.search(
                        r"(?<=\/Projetos\/).*?(?=\/Arquivos\ do\ Projeto)",
                        caminho_projeto[0],
                    )
                    projeto_arquivo = projeto_arquivo.group(0)
                    print("-----------------------------------------------")
                    print("projeto_arquivo:", projeto_arquivo)
                    print("-----------------------------------------------")

                    pasta_destino = os.path.join(
                        diretorio_raiz,
                        projeto_arquivo,
                        "Arquivos do Projeto",
                        "Para Entrega",
                        "Aprovados",
                    )

                    # Itera sobre os resultados e move cada arquivo para a pasta de destino
                    conn = sqlite3.connect("database.db")
                    c = conn.cursor()
                    for i in range(tamanho):
                        nome_do_arquivo = os.path.basename(caminho_projeto[i])
                        novo_caminho = os.path.join(pasta_destino, nome_arq[i])

                        # Atualiza o status do arquivo na base de dados
                        novo_status = status[
                            (status.index(status_atual[i]) + 1) % len(status)
                        ]
                        query = f"""UPDATE arquivos SET caminho='{novo_caminho}', status = '{novo_status}', responsavel = '{username}', data_aprovado = '{data_atualizada}' WHERE id = '{id_documentos[i]}' """
                        c.execute(query)

                        c.execute(
                            "INSERT INTO log_tarefas (nome, projeto, status, data_status, responsavel) VALUES (?, ?, ?, ?, ?)",
                            (
                                nome_do_arquivo,
                                projeto,
                                status_atual[0],
                                data_atualizada,
                                username,
                            ),
                        )
                        conn.commit()

                        conn.commit()

                        shutil.move(caminho_projeto[i], pasta_destino)

                    conn.close()
                    linha_selecionada = []
                    df_tabela = df_tabela.dropna()
        elif status_atualizar == "gerar_grd":
            control = "renomear"

        elif status_atualizar == "cancelar":
            linha_selecionada = []
            df_tabela = df_tabela.dropna()
            df_selecionado = df_selecionado.dropna()
            control = ""

        return redirect(
            url_for(
                "user_projetos",
                projeto=projeto,
                atualizar=atualizar,
                pasta_destino=pasta_destino,
                data_atualizada=data_atualizada,
                abreviacao=abreviacao,
            )
        )


@app.route("/renomear_pasta", methods=["POST"])
def renomear_pasta():
    if "username" in session:
        global df_tabela
        global pasta_destino
        global novo_caminho
        global atualizar
        global control
        global mudar_status
        global linha_selecionada
        global df_selecionado

        username = session["username"]
        projeto = request.form["projeto"]
        nome_pasta = request.form["nome_pasta"]
        tamanho = len(linha_selecionada)

        conn = sqlite3.connect("database.db")
        query = "SELECT * FROM arquivos"
        df_tabela = pd.read_sql_query(query, conn)
        df_selecionado = df_tabela.loc[df_tabela["id"].isin(linha_selecionada)]
        nome_arq = df_selecionado["nome"].values.tolist()
        caminho_projeto_df = df_selecionado["caminho"].values.tolist()
        status_atual = df_selecionado["status"].values.tolist()
        id_documentos = df_selecionado["id"].values.tolist()
        conn.close()

        caminho_projeto = re.search(
            r"(?<=).*?(?=\/Arquivos\ do\ Projeto)", caminho_projeto_df[0]
        )
        caminho_projeto = caminho_projeto.group(0)

        projeto_arquivo = re.search(
            r"(?<=\/Projetos\/).*?(?=\/Arquivos\ do\ Projeto)", caminho_projeto_df[0]
        )
        projeto_arquivo = projeto_arquivo.group(0)

        # -------------------- Buscar nome do projeto na pasta --------------------

        # caminho_projeto = os.path.join(diretorio_raiz, projeto[0])
        caminho_verificado = os.path.join(
            caminho_projeto, "Arquivos do Projeto", "Para Entrega"
        )

        now = datetime.datetime.now()
        data_envio = now.strftime("%d/%m/%Y")

        conn = sqlite3.connect("database.db")
        query = "SELECT * FROM arquivos"
        df_renomear = pd.read_sql_query(query, conn)
        conn.close()

        caminho_atual = os.path.join(
            diretorio_raiz,
            projeto_arquivo,
            "Arquivos do Projeto",
            "Para Entrega",
            "Aprovados",
        )

        novo_caminho = os.path.join(
            diretorio_raiz,
            projeto_arquivo,
            "Arquivos do Projeto",
            "Para Entrega",
            nome_pasta,
        )

        os.makedirs(novo_caminho, exist_ok=True)

        conn = sqlite3.connect("database.db")
        c = conn.cursor()

        for index, row in df_selecionado.iterrows():
            nome_do_arquivo = os.path.basename(row["caminho"])
            caminho_arq = os.path.join(caminho_atual, nome_do_arquivo)

            shutil.move(caminho_arq, novo_caminho)

            if row["caminho"] == caminho_arq:
                caminho_arq = os.path.join(novo_caminho, nome_do_arquivo)
                query = f"""UPDATE arquivos SET caminho = '{caminho_arq}', status = 'Entregue', data_entregue = '{data_atualizada}' WHERE id = {row["id"]}"""
                c.execute(query)

                c.execute(
                    "INSERT INTO log_tarefas (nome, projeto, status, data_status, responsavel) VALUES (?, ?, ?, ?, ?)",
                    (
                        nome_do_arquivo,
                        projeto_arquivo,
                        "Para Entrega",
                        data_atualizada,
                        username,
                    ),
                )
                conn.commit()

                conn.commit()
        conn.close()

        control = "gerar_grd"

        return redirect(
            url_for(
                "user_projetos",
                projeto=projeto,
                atualizar=atualizar,
                data_envio=data_envio,
            )
        )


@app.route("/gerar_grd", methods=["POST"])
def gerar_grd():
    if "username" in session:
        global abreviacao
        global diretorio_raiz
        global pasta_destino
        global linha_selecionada
        global novo_caminho
        global df_projeto
        global df_arquivo
        global atualizar
        global control
        global df_tabela
        global mudar_status
        global caminho_padrao

        tipos = {
            "A": "PRELIMINAR",
            "B": "PARA APROVAÇAO",
            "C": "PARA CONHECIMENTO",
            "D": "PARA COTAÇAO",
            "E": "PARA CONSTRUÇAO",
            "F": "CONFORME COMPRADO",
            "G": "CONFORME CONSTRUIDO",
            "H": "CANCELADO",
        }
        username = session["username"]
        projeto_atual = request.form["projeto"]
        tipo = request.form["tipo"]
        descricao = request.form["descricao"]
        verificado = request.form["verificado"]
        aprovado = request.form["aprovado"]
        autorizado = request.form["autorizado"]

        projeto = projeto_atual.strip()

        conn = sqlite3.connect("database.db")
        query = "SELECT * FROM arquivos"
        df_tabela = pd.read_sql_query(query, conn)
        df_selecionado = df_tabela.loc[df_tabela["id"].isin(linha_selecionada)]
        nome_arq = df_selecionado["nome"].values.tolist()
        caminho_projeto_df = df_selecionado["caminho"].values.tolist()
        status_atual = df_selecionado["status"].values.tolist()
        id_documentos = df_selecionado["id"].values.tolist()
        conn.close()

        caminho_projeto = re.search(
            r"(?<=).*?(?=\/Arquivos\ do\ Projeto)", caminho_projeto_df[0]
        )
        caminho_projeto = caminho_projeto.group(0)

        projeto_arquivo = re.search(
            r"(?<=\/Projetos\/).*?(?=\/Arquivos\ do\ Projeto)", caminho_projeto_df[0]
        )
        projeto_arquivo = projeto_arquivo.group(0)

        # -------------------- Buscar nome do projeto na pasta --------------------
        caminho_verificado = os.path.join(
            caminho_projeto, "Arquivos do Projeto", "Para Entrega"
        )

        # buscar abreviacao
        conn = sqlite3.connect("database.db")
        query = "SELECT * FROM dados_projeto"
        df_projeto = pd.read_sql_query(query, conn)

        result_projeto = df_projeto[df_projeto["projeto"] == projeto]

        abreviacao_empresa = result_projeto["abreviacao"].iloc[0]
        descricao_projeto = result_projeto["descricao"].iloc[0]

        # ------------------- CAMINHOS PADROES -------------------------

        caminho_ld_padrao = os.path.join(caminho_padrao, "LD_padrao.xlsx")
        caminho_grd_padrao = os.path.join(caminho_padrao, "GRD_padrao.xlsx")

        pastas = []
        pasta_GRD_recente = None
        pasta_ld_recente = None
        ultima_data_criacao = 0
        contador_GRD = 0
        # Percorre todos os diretórios e arquivos no caminho fornecido
        for diretorio_atual, subdiretorios, arquivos in os.walk(caminho_verificado):
            subdiretorios.sort()
            for subdiretorio in subdiretorios:
                pastas.append(subdiretorio)
                # Verifica se o nome da pasta contém "GRD"
                if re.search(r"GRD", subdiretorio):
                    contador_GRD += 1
                    caminho_pasta = os.path.join(diretorio_atual, subdiretorio)
                    # Verifica se é a pasta mais recente
                    if contador_GRD != 1:
                        pasta_GRD_anterior = pasta_GRD_recente
                        pasta_GRD_recente = caminho_pasta
                    else:
                        pasta_GRD_anterior = None
                        pasta_GRD_recente = caminho_pasta
        # Gerar LD
        if pasta_GRD_anterior == None:
            # Criaçao de uma nova GRD, primeira entrega
            # Abrir o arquivo Excel
            workbook = load_workbook(caminho_ld_padrao)

            # Obter a planilha desejada]
            planilha = workbook["F. Rosto"]

            # Nome da Empresa-Cidade
            planilha["J1"].value = projeto

            """shape_empresa = planilha.shapes["nome_empresa1"]
            shape_empresa.text = projeto  # projeto  ----------------"""
            
            # Descriçao do projeto
            planilha["A5"].value = descricao_projeto
            # Logica para adicionar a primeira revisao
            X = 12  # linha 12 é a primeira a ser adicionada a revisao 0
            planilha["A" + str(X)].value = "0"
            # Tipo (TE)
            planilha["B" + str(X)].value = str(tipo)
            # Descriçao
            planilha["C" + str(X)].value = str(descricao)
            # Feito por:
            planilha["J" + str(X)].value = abreviacao
            # Verificado
            planilha["K" + str(X)].value = verificado
            # Aprovado
            planilha["L" + str(X)].value = aprovado
            # Aut.
            planilha["M" + str(X)].value = autorizado

            # Data de envio
            now = datetime.datetime.now()
            data_envio = now.strftime("%d/%m/%Y")
            planilha["N" + str(X)].value = data_envio

            nome_arq_ld = "ABS-" + abreviacao_empresa + "-LD-" + "001" + "_R0" + ".xlsx"
            planilha["J6"].value = nome_arq_ld
            nome_ld = os.path.join(novo_caminho, nome_arq_ld)

            # ------------------- Modificando a planilha LISTA -------------------
            planilha = workbook["Lista"]

            conn = sqlite3.connect("database.db")
            query = "SELECT * FROM arquivos"
            df_tabela = pd.read_sql_query(query, conn)
            df_selecionado = df_tabela.loc[df_tabela["id"].isin(linha_selecionada)]
            nome_arq = df_selecionado["nome"].values.tolist()
            conn.close()

            X = 11
            planilha["B" + str(X)].value = nome_arq_ld
            planilha["C" + str(X)].value = "LISTA DE DOCUMENTOS"
            planilha["D" + str(X)].value = "0"
            planilha["F" + str(X)].value = "0"
            planilha["I" + str(X)].value = "LD"
            planilha["J" + str(X)].value = "2"
            planilha["H" + str(X)].value = "A4"
            planilha["K" + str(X)].value = "100"
            planilha["L" + str(X)].value = "=1/8*J11"

            padrao = r"R\d+"

            X = 12

            conn = sqlite3.connect("database.db")
            query = "SELECT * FROM dados_arquivo"
            df_arquivo = pd.read_sql_query(query, conn)

            for nome in nome_arq:
                result_arquivo = df_arquivo[df_arquivo["nome"] == nome]
                titulo = result_arquivo["titulo"].iloc[0]
                planilha["B" + str(X)].value = nome
                planilha["C" + str(X)].value = str(titulo)

                resultado = re.search(r"_R(\d+)", nome)
                if resultado:
                    revisao = resultado.group(1)
                    planilha["D" + str(X)].value = revisao
                expressoes = [
                    "LC",
                    "LD",
                    "LE",
                    "LI",
                    "LM",
                    "PLT",
                    "ET",
                    "SE",
                    "ILU",
                    "UNF",
                    "SPDA",
                    "TRIF",
                    "REDE",
                    "ARQR",
                    "DIAG",
                ]
                padrao = r"\b(?:{})\b".format("|".join(expressoes))
                match = re.search(padrao, nome)
                if match:
                    tipo_expressao = match.group(0)
                planilha["I" + str(X)].value = tipo_expressao
                planilha["F" + str(X)].value = "0"
                planilha["J" + str(X)].value = "1"
                planilha["K" + str(X)].value = "100"
                planilha["L" + str(X)].value = "=1/8*J" + str(X)

                A1 = ["SPDA", "SE", "UNF", "TRIF", "PLT", "SPDA", "REDE", "ILU"]
                A2 = []
                A3 = ["DIAG"]
                A4 = [
                    "LC",
                    "LD",
                    "LE",
                    "LI",
                    "LM",
                    "ET",
                ]

                if tipo_expressao in A1:
                    planilha["H" + str(X)].value = "A1"
                if tipo_expressao in A2:
                    planilha["H" + str(X)].value = "A2"
                if tipo_expressao in A3:
                    planilha["H" + str(X)].value = "A3"
                if tipo_expressao in A4:
                    planilha["H" + str(X)].value = "A4"
                X = X + 1

            workbook.save(nome_ld)
            # Fechar o arquivo Excel
            workbook.close()

            seq_grd = "01"
            nome_grd_parte = "GRD-ABS-" + abreviacao_empresa + "-0" + seq_grd + ".xlsx"
            nome_grd = os.path.join(novo_caminho, nome_grd_parte)

        else:
            # Atualizaçao de uma GRD ja existente, subir revisao

            for diretorio_atual, subdiretorios, arquivos in os.walk(pasta_GRD_anterior):
                for arquivo in arquivos:
                    if re.search(r"LD", arquivo):
                        arquivo_ld_anterior = arquivo
                        partes = arquivo_ld_anterior.split("_R")
                        parte1 = arquivo_ld_anterior.split("_R")[1]
                        parte2 = parte1.rsplit(".", 1)[0]
                        if len(partes) > 1:
                            nova_revisao_ld = int(parte2) + 1
                            inicio_nome_ld = partes[0]
                            nome_arq_ld = (
                                inicio_nome_ld + "_R" + str(nova_revisao_ld) + ".xlsx"
                            )

                        caminho_ld_anterior = os.path.join(
                            diretorio_atual, arquivo_ld_anterior)

                    if re.search(r"GRD", arquivo):
                        arquivo_grd_anterior = arquivo
                        partes = arquivo_grd_anterior.split("-")
                        if len(partes) > 2:
                            nova_revisao_grd = int(partes[-1].split(".")[0]) + 1
                            inicio_nome_grd = partes[0] + partes[1] + partes[2]
                            novo_nome_grd = (
                                inicio_nome_grd + str(nova_revisao_grd) + ".xlsx"
                            )

                        caminho_grd_anterior = os.path.join(
                            diretorio_atual, arquivo_grd_anterior
                        )

            # Abrir o arquivo Excel
            workbook = load_workbook(caminho_ld_anterior)
            # Obter a planilha desejada
            planilha = workbook["F. Rosto"]

            planilha["J6"].value = nome_arq_ld

            # Logica para ler qual a ultima revisao
            proxima_linha = None
            for linha in range(12, 31 + 1):
                if planilha.cell(row=linha, column=1).value is None:
                    X = linha
                    ultima_revisao = X - 13
                    break

            # Buscar o nome de todos os arquivos a serem enviados na pagina
            planilha["A" + str(X)].value = str(ultima_revisao + 1) 
            planilha["B" + str(X)].value = str(tipo)
            planilha["C" + str(X)].value = str(descricao)
            planilha["J" + str(X)].value = abreviacao
            planilha["K" + str(X)].value = verificado
            planilha["L" + str(X)].value = aprovado
            planilha["M" + str(X)].value = autorizado

            # Data de envio
            now = datetime.datetime.now()
            data_envio = now.strftime("%d/%m/%Y")
            planilha["N" + str(X)].value = data_envio

            nome_ld = os.path.join(caminho_pasta, nome_arq_ld)

            # ------------------- Modificando a planilha LISTA -------------------
            planilha = workbook["Lista"]

            conn = sqlite3.connect("database.db")
            query = "SELECT * FROM arquivos"
            df_tabela = pd.read_sql_query(query, conn)
            df_selecionado = df_tabela.loc[df_tabela["id"].isin(linha_selecionada)]
            nome_arq = df_selecionado["nome"].values.tolist()
            conn.close()

            padrao = r"R\d+"

            X = 11
            
            planilha["B" + str(X)].value = nome_arq_ld
            planilha["F" + str(X)].value = str(nova_revisao_ld)

            conn = sqlite3.connect("database.db")
            query = "SELECT * FROM dados_arquivo"
            df_arquivo = pd.read_sql_query(query, conn)


            proxima_linha = None
            for linha in range(12, 31 + 1):
                if planilha.cell(row=linha, column=2).value is None:
                    X = linha
                    break

            # Logica para ler qual a ultima revisao
            for nome in nome_arq:
                result_arquivo = df_arquivo[df_arquivo["nome"] == nome]
                titulo = result_arquivo["titulo"].iloc[0]
                planilha["B" + str(X)].value = nome
                planilha["C" + str(X)].value = str(titulo)
                planilha["D" + str(X)].value = "0"

                resultado = re.search(r"_R(\d+)", nome)
                if resultado:
                    revisao = resultado.group(1)
                    planilha["F" + str(X)].value = revisao
                expressoes = [
                    "LC",
                    "LE",
                    "LI",
                    "LM",
                    "PLT",
                    "ET",
                    "SE",
                    "ILU",
                    "UNF",
                    "SPDA",
                    "TRIF",
                    "REDE",
                    "ARQR",
                    "DIAG",
                ]
                padrao = r"\b(?:{})\b".format("|".join(expressoes))
                match = re.search(padrao, nome)
                if match:
                    tipo_expressao = match.group(0)
                planilha["I" + str(X)].value = tipo_expressao
                planilha["J" + str(X)].value = "1"
                planilha["K" + str(X)].value = "100"
                planilha["L" + str(X)].value = "=1/8*J" + str(X)

                A1 = ["SPDA", "SE", "UNF", "TRIF", "PLT", "SPDA", "REDE", "ILU"]
                A2 = []
                A3 = ["DIAG"]
                A4 = ["LC", "LE", "LI", "LM", "ET"]

                if tipo_expressao in A1:
                    planilha["H" + str(X)].value = "A1"
                if tipo_expressao in A2:
                    planilha["H" + str(X)].value = "A2"
                if tipo_expressao in A3:
                    planilha["H" + str(X)].value = "A3"
                if tipo_expressao in A4:
                    planilha["H" + str(X)].value = "A4"
                X = X + 1

            workbook.save(nome_ld)
            # Fechar o arquivo Excel
            workbook.close()

            if pasta_GRD_anterior is not None:
                for diretorio_atual, subdiretorios, arquivos in os.walk(
                    pasta_GRD_anterior
                ):
                    for arquivo in arquivos:
                        if re.search(r"GRD", arquivo):
                            parte_split_grd = arquivo.split("-")
                            seq_grd = parte_split_grd[3].split(".")[0]

                            if int(seq_grd) > 9:
                                seq_grd = int(seq_grd) + 1
                                nome_grd_parte = (
                                    "GRD-ABS-"
                                    + abreviacao_empresa
                                    + "-0"
                                    + str(seq_grd)
                                    + ".xlsx"
                                )
                                nome_grd = os.path.join(novo_caminho, nome_grd_parte)
                            else:
                                seq_grd = "0" + str(int(seq_grd) + 1)
                                nome_grd_parte = (
                                    "GRD-ABS-"
                                    + abreviacao_empresa
                                    + "-0"
                                    + seq_grd
                                    + ".xlsx"
                                )
                                nome_grd = os.path.join(novo_caminho, nome_grd_parte)

        # ------------------------- GRD ------------------------------------
        workbook = load_workbook(caminho_grd_padrao)
        planilha = workbook["GRD"]

        planilha["N7"].value = data_envio
        parte_num_grd = nome_grd.split("/")
        numero_grd = (parte_num_grd[9].split("-")[3]).split(".")[0]
        planilha["N10"].value = numero_grd

        parte_split_ld = nome_arq_ld.split("_R", 1)
        nova_revisao = parte_split_ld[1].split(".")[0]

        X = 36

        planilha["A" + str(X)].value = "01"
        planilha["H" + str(X)].value = nome_arq_ld
        planilha["K" + str(X)].value = "Lista de documentos"
        planilha["O" + str(X)].value = nova_revisao
        planilha["P" + str(X)].value = str(tipo)
        planilha["Q" + str(X)].value = "A4"
        planilha["S" + str(X)].value = "1"
        planilha["T" + str(X)].value = "DOC.ELETRÔNICO"
        planilha["V" + str(X)].value = "Original"

        padrao = r"R\d+"
        X = 37

        conn = sqlite3.connect("database.db")
        query = "SELECT * FROM dados_arquivo"
        df_arquivo = pd.read_sql_query(query, conn)

        for nome in nome_arq:
            planilha["A" + str(X)].value = str(X - 36)
            result_arquivo = df_arquivo[df_arquivo["nome"] == nome]
            titulo = result_arquivo["titulo"].iloc[0]
            planilha["H" + str(X)].value = nome
            planilha["K" + str(X)].value = str(titulo)

            resultado = re.search(r"_R(\d+)", nome)
            if resultado:
                revisao = resultado.group(1)
                planilha["O" + str(X)].value = revisao
            planilha["P" + str(X)].value = str(tipo)
            planilha["S" + str(X)].value = "1"
            planilha["T" + str(X)].value = "DOC.ELETRÔNICO"
            planilha["V" + str(X)].value = "Original"

            if tipo_expressao in A1:
                planilha["Q" + str(X)].value = "A1"
            if tipo_expressao in A2:
                planilha["Q" + str(X)].value = "A2"
            if tipo_expressao in A3:
                planilha["Q" + str(X)].value = "A3"
            if tipo_expressao in A4:
                planilha["Q" + str(X)].value = "A4"
            X = X + 1

        workbook.save(nome_grd)
        # Fechar o arquivo Excel
        workbook.close()

        pasta_GRD_recente = None
        pasta_GRD_anterior = None

        control = ""

        return redirect(
            url_for(
                "user_projetos",
                projeto=projeto,
            )
        )


# -------------------------------------- CRIAR ARQUIVOS --------------------------------------


@app.route("/criar_arquivo", methods=["POST"])
def criar_arquivo():
    if "username" in session:
        global df_tabela
        global diretorio_raiz
        global diretorio_default
        global df_projeto
        global df_arquivo

        projeto_recebido = request.form["projeto"]

        username = session["username"]
        disciplina = request.form["disc"]
        sub = request.form["sub"]
        titulo = request.form["titulo"]

        now = datetime.datetime.now()
        data_atualizada = now.strftime("%d/%m/%Y %H:%M:%S")

        diretorio_novos = os.path.join(
            diretorio_raiz,
            difflib.get_close_matches(projeto_recebido, os.listdir(diretorio_raiz))[0],
        )

        # -------------------- Buscar nome do projeto na pasta --------------------
        folders = [
            f
            for f in os.listdir(diretorio_raiz)
            if os.path.isdir(os.path.join(diretorio_raiz, f))
        ]

        # Criar uma expressão regular para verificar se o nome da pasta contém a parte fornecida
        pattern = re.compile(rf".*{re.escape(projeto_recebido)}.*", re.IGNORECASE)

        # Percorrer os nomes das pastas e verificar se eles correspondem à expressão regular
        proj_prox = [f for f in folders if re.match(pattern, f)]

        # ---------------------Buscar abreviacao da empresa-------------------

        # Buscar o titulo
        conn = sqlite3.connect("database.db")
        query = "SELECT * FROM dados_projeto"
        df_projeto = pd.read_sql_query(query, conn)

        proj_espaco = projeto_recebido.lstrip()

        result = df_projeto.loc[df_projeto["projeto"] == proj_espaco]
        abreviacao_empresa = result.loc[0, "abreviacao"]

        arquivo_existente = request.form["arquivo_existente"]
        name, extension = os.path.splitext(arquivo_existente)

        # --------------- Sequencia dos arquivos ---------------------
        caminho_projeto = os.path.join(diretorio_raiz, proj_prox[0])
        df_teste = pd.DataFrame(columns=["nome", "projeto"])

        # ------------------- Buscar os dados no CSV ---------------------
        disc = []
        with open("disc.csv", "r", encoding="utf-8-sig") as arquivo:
            for line in arquivo:
                fields = line.strip().split(";")
                if fields[1] == disciplina:
                    disciplina = fields[0]
        sub1 = []
        with open("sub.csv", "r", encoding="utf-8-sig") as arquivo:
            for line in arquivo:
                fields = line.strip().split(";")
                if fields[1] == sub:
                    sub = fields[0]

        nome_arquivo = (
            "ABS-"
            + str(abreviacao_empresa)
            + "-"
            + str(disciplina)
            + "-"
            + str(sub)
            + "-"
            + "0"
            + "_R0"
            + str(extension)
        )

        nome_arq = None
        conn = sqlite3.connect("database.db")
        query = "SELECT * FROM arquivos"
        df_teste = pd.read_sql_query(query, conn)

        partes = nome_arquivo.split("-", 4)
        padrao = "-".join(partes[:4])

        result_teste = df_teste[df_teste["nome"].str.contains(padrao, regex=False)]

        if not result_teste.empty:
            sequenciais = []
            for nome in result_teste["nome"]:
                parte = nome.split("_")
                numero_parte = parte[0]

                partes = numero_parte.split("-")
                sequencial = partes[4]
                # sequencial = numero_parte.split("_")[0]
                sequenciais.append(sequencial)
            if sequenciais:
                ultimo_sequencial = max(sequenciais)
                novo_sequencial = int(ultimo_sequencial) + 1
                nome_arquivo = (
                    "ABS-"
                    + str(abreviacao_empresa)
                    + "-"
                    + str(disciplina)
                    + "-"
                    + str(sub)
                    + "-"
                    + str(novo_sequencial)
                    + "_R0"
                    + str(extension)
                )

        caminho_origem = os.path.join(diretorio_default, arquivo_existente)

        # Caminho de destino tem que ser a pasta do usuario na area de trabalho
        caminho_destino = os.path.join(
            diretorio_novos,
            "Arquivos do Projeto",
            "Area de Trabalho",
            username,
            nome_arquivo,
        )

        shutil.copyfile(caminho_origem, caminho_destino)

        projeto_arquivo = re.search(r"\d...([A-Za-z\s]+[\w-]+)", caminho_destino)
        projeto_arquivo = projeto_arquivo.group(1) if proj_prox else None

        conn = sqlite3.connect("database.db")
        c = conn.cursor()
        c.execute(
            "INSERT INTO arquivos (nome, status, responsavel, data_criado, caminho, projeto) VALUES (?, ?, ?, ?, ?, ?)",
            (
                nome_arquivo,
                "Criado",
                username,
                data_atualizada,
                caminho_destino,
                projeto_arquivo,
            ),
        )

        c.execute(
            "INSERT INTO dados_arquivo (nome, projeto, titulo) VALUES (?, ?, ?)",
            (nome_arquivo, proj_prox[0], titulo),
        )

        c.execute(
            "INSERT INTO log_tarefas (nome, projeto, status, data_status, responsavel) VALUES (?, ?, ?, ?, ?)",
            (
                nome_arquivo,
                projeto_recebido,
                "Criado",
                data_atualizada,
                username,
            ),
        )
        conn.commit()
        conn.commit()
        conn.close()

        return redirect(
            url_for(
                "documentos",
                projeto=projeto_recebido,
                titulo=titulo,
            )
        )


@app.route("/criar_projeto", methods=["POST"])
def criar_projeto():
    if "username" in session:
        global diretorio_raiz
        global df_projeto
        global pasta_padrao_projeto

        print("---------------- CRIAR PROJETO ------------------------")

        nome_projeto_inicial = request.form["nome_projeto"]
        print("nome_projeto_inicial: ", nome_projeto_inicial)
        print(nome_projeto_inicial)
        abreviacao_empresa = request.form["abreviacao_empresa"]
        print("abreviacao_empresa: ", abreviacao_empresa)
        descricao_projeto = request.form["descricao_projeto"]
        print("descricao_projeto: ", descricao_projeto)

        projetos = os.listdir(diretorio_raiz)
        maior_numero = 0
        for projeto in projetos:
            numero = re.search(r"^(\d+) -", projeto)
            if numero:
                numero_int = int(numero.group(1))
                if numero_int > maior_numero:
                    maior_numero = numero_int
        numero_projeto = maior_numero + 1
        nome_projeto = str(numero_projeto) + " - " + nome_projeto_inicial

        print("nome_projeto: ", nome_projeto)
        pasta_atualizada = os.path.join(diretorio_raiz, nome_projeto)

        # adicionar o nome do projeto, a abreviacao e a descricao do projeto

        conn = sqlite3.connect("database.db")
        c = conn.cursor()
        query = f"""INSERT INTO dados_projeto (projeto, abreviacao, descricao) VALUES ('{nome_projeto_inicial}', '{abreviacao_empresa}', '{descricao_projeto}')"""
        c.execute(query)
        conn.commit()
        conn.close()

        shutil.copytree(pasta_padrao_projeto, pasta_atualizada)
        return redirect(
            url_for(
                "projetos",
            )
        )


@app.route("/configuracoes", methods=["GET", "POST"])
def configuracoes():
    global diretorio_raiz
    global caminho_padrao
    global diretorio_default
    global pasta_padrao_projeto

    if request.method == "POST":
        config_valor = request.form.get("config_valor", None)

        if config_valor == "config_diretorios":
            diretorio_projetos = request.form.get("diretorio_projetos", None)
            diretorio_grd = request.form.get("diretorio_grd", None)
            diretorio_padrao = request.form.get("diretorio_padrao", None)
            pasta_projeto = request.form.get("pasta_projeto", None)

            if diretorio_projetos is not None:
                diretorio_raiz = diretorio_projetos
            if diretorio_grd is not None:
                caminho_padrao = diretorio_grd
            if diretorio_padrao is not None:
                diretorio_default = diretorio_padrao
            if pasta_projeto is not None:
                pasta_padrao_projeto = pasta_projeto

    # Renderizar o template da página de configurações
    return render_template("configuracoes.html")


# --------------------------------------------------------------------------------

if __name__ == "__main__":
    app.run(debug=True, host="0.0.0.0")
